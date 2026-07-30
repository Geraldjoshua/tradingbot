"""
Distill the OptionStrat master workbooks into ONE tiny JSON the trading app reads.

WHY THIS EXISTS
The app only ever needs the per-ticker Aggregate numbers — but reading them out
of a master .xlsx costs ~8-9 SECONDS once the Active sheet grows to ~150k rows,
because openpyxl must parse the workbook's shared-string table even in read_only
mode and even if you only touch one sheet. Multiply by every ticker the
auto-trader evaluates and flow reads dominate the whole cycle.

This script pays that cost ONCE, locally, and writes `flow_cache.json`:

    { "generated": "2026-07-29T14:03:11",
      "tickers": { "TSLA": {"bull":4.2e6,"bear":9e5,"u":true,"k":false,
                            "cw":[285,290,300],"pw":[250,240,230]}, ... } }

Typical size: tens of KB for a few thousand tickers (vs 8.7 MB of workbooks).

TWO PAYOFFS
  1. Reads drop from ~8.7s to ~milliseconds.
  2. Deploying flow data to a cloud host means copying ONE small file instead of
     three multi-MB workbooks.

USAGE — run right after optionstrat_master_builder.py, in the same folder:
    python flow/build_flow_cache.py [dir]

Or chain them:
    python optionstrat_master_builder.py && python flow/build_flow_cache.py
"""

import json
import sys
from datetime import datetime
from pathlib import Path

MASTERS = {
    "live":    "flow_master.xlsx",
    "unusual": "flow_unusual_master.xlsx",
    "knows":   "flow_knows_master.xlsx",
}
CACHE_NAME = "flow_cache.json"


def parse_premium(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lstrip("$").lower()
    mult = 1.0
    if s.endswith("m"):
        mult, s = 1e6, s[:-1]
    elif s.endswith("k"):
        mult, s = 1e3, s[:-1]
    elif s.endswith("b"):
        mult, s = 1e9, s[:-1]
    try:
        return float(s) * mult
    except ValueError:
        return 0.0


def read_aggregate(path):
    """{ticker: (bull, bear, call_walls, put_walls)} from a master's Aggregate."""
    out = {}
    if not path.exists():
        return out
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Aggregate" not in wb.sheetnames:
        wb.close()
        return out
    rows = wb["Aggregate"].iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        wb.close()
        return out
    idx = {str(h): i for i, h in enumerate(header)}
    ti = idx.get("ticker", 0)
    bi, ri = idx.get("bullish_premium"), idx.get("bearish_premium")

    def walls(row, pfx):
        vals = []
        for n in (1, 2, 3):
            c = idx.get(f"{pfx}_{n}")
            if c is not None and row[c] not in (None, ""):
                try:
                    vals.append(float(row[c]))
                except (TypeError, ValueError):
                    pass
        return vals

    for row in rows:
        if not row or row[ti] in (None, ""):
            continue
        t = str(row[ti]).upper()
        out[t] = (
            parse_premium(row[bi]) if bi is not None else 0.0,
            parse_premium(row[ri]) if ri is not None else 0.0,
            walls(row, "call_wall"),
            walls(row, "put_wall"),
        )
    wb.close()
    return out


def build(directory="."):
    d = Path(directory)
    t0 = datetime.now()
    live = read_aggregate(d / MASTERS["live"])
    unusual = set(read_aggregate(d / MASTERS["unusual"]).keys())
    knows = set(read_aggregate(d / MASTERS["knows"]).keys())

    tickers = {}
    for t, (bull, bear, cw, pw) in live.items():
        tickers[t] = {
            "bull": round(bull, 2), "bear": round(bear, 2),
            "u": t in unusual, "k": t in knows,
            "cw": cw, "pw": pw,
        }
    # Names present only in the higher-signal books still deserve an entry.
    for t in (unusual | knows):
        if t not in tickers:
            tickers[t] = {"bull": 0.0, "bear": 0.0, "u": t in unusual, "k": t in knows, "cw": [], "pw": []}

    blob = {
        "generated": t0.isoformat(timespec="seconds"),
        "sources": [m for m, f in MASTERS.items() if (d / MASTERS[m]).exists()],
        "count": len(tickers),
        "tickers": tickers,
    }
    out = d / CACHE_NAME
    out.write_text(json.dumps(blob, separators=(",", ":")))
    return out, blob


def main():
    directory = sys.argv[1] if len(sys.argv) > 1 else "."
    try:
        out, blob = build(directory)
        kb = out.stat().st_size / 1024
        print(f"wrote {out}  ({blob['count']} tickers, {kb:.1f} KB, sources: {', '.join(blob['sources']) or 'none'})")
    except Exception as e:
        print(f"failed: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
