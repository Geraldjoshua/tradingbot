"""
OptionStrat flow-conviction reader  (consumed by the Node auto-trader).

Given a ticker, this reads the master workbooks the builder already produces
(flow_master.xlsx / flow_unusual_master.xlsx / flow_knows_master.xlsx) and returns
a single conviction verdict as JSON on stdout:

    python flow/optionstrat_flow.py TSLA /path/to/optionstrat_dir

    {
      "ticker": "TSLA",
      "found": true,
      "bullish_premium": 4210000.0,
      "bearish_premium": 900000.0,
      "net_premium": 3310000.0,
      "direction": "bullish",          # bullish | bearish | neutral
      "score": 0.82,                   # 0..1 conviction strength (|net| / total)
      "in_unusual": true,              # ticker also present in the "Highly Unusual" feed
      "in_knows": false,               # ... and the "In The Know" feed
      "call_walls": [285.0, 290.0, 300.0],
      "put_walls":  [250.0, 240.0, 230.0],
      "sources": ["flow_master.xlsx", "flow_unusual_master.xlsx"]
    }

Design notes
------------
* The "live" master (flow_master.xlsx) is the broad book and supplies the
  bullish/bearish premium used for direction + score. Presence in the *unusual*
  and *knows* feeds is surfaced as boolean boosters (the Node side can weight
  them). This mirrors how the builder splits the three feeds.
* If no master workbook exists yet, we fall back to today's day-CSVs
  (flow_YYYY-MM-DD.csv etc.) and aggregate sentiment premium directly, so the
  conviction signal works even before the first master build has run.
* Never raises: any failure returns {"found": false, "error": ...} so the trading
  loop degrades to "no OptionStrat signal" instead of crashing.

The Aggregate sheet header (fixed, from optionstrat_master_builder.py) is:
  ticker, bullish_premium, bearish_premium,
  call_wall_1, call_wall_2, call_wall_3, put_wall_1, put_wall_2, put_wall_3
"""

import csv
import json
import sys
from datetime import datetime
from pathlib import Path

BULLISH_SENTS = {"bullish", "very-bullish"}
BEARISH_SENTS = {"bearish", "very-bearish"}

MASTERS = {
    "live":    "flow_master.xlsx",
    "unusual": "flow_unusual_master.xlsx",
    "knows":   "flow_knows_master.xlsx",
}


def parse_premium(premium_str, tag=None):
    """'$1.16m' -> 1160000.0 (same rules as the master builder)."""
    if premium_str is None:
        return None
    s = str(premium_str).strip().lstrip("$").lower()
    mult = 1.0
    if s.endswith("m"):
        mult, s = 1_000_000.0, s[:-1]
    elif s.endswith("k"):
        mult, s = 1_000.0, s[:-1]
    elif s.endswith("b"):
        mult, s = 1_000_000_000.0, s[:-1]
    elif tag:
        t = str(tag).strip().lower()
        mult = {"m": 1e6, "k": 1e3, "b": 1e9}.get(t, 1.0)
    try:
        return float(s) * mult
    except ValueError:
        return None


# ── FAST PATH: flow_cache.json (built by flow/build_flow_cache.py) ──────────
# Reading the Aggregate sheet out of an 8MB master costs ~8.7s once Active grows
# to ~150k rows (openpyxl must parse the shared-string table regardless of
# read_only). The cache turns that into a few milliseconds. We always prefer it
# and fall back to the workbooks if it's missing.
CACHE_NAME = "flow_cache.json"
_CACHE = {"path": None, "blob": None}


def _load_cache(directory):
    path = Path(directory) / CACHE_NAME
    if _CACHE["path"] == str(path) and _CACHE["blob"] is not None:
        return _CACHE["blob"]
    if not path.exists():
        return None
    try:
        blob = json.loads(path.read_text())
    except Exception:
        return None
    if not isinstance(blob, dict) or "tickers" not in blob:
        return None
    _CACHE["path"], _CACHE["blob"] = str(path), blob
    return blob


def _agg_from_cache(directory, ticker):
    blob = _load_cache(directory)
    if not blob:
        return None
    e = blob["tickers"].get(ticker)
    if not e:
        return None
    return (e.get("bull", 0.0), e.get("bear", 0.0), e.get("cw", []), e.get("pw", [])), \
        bool(e.get("u")), bool(e.get("k"))


def _agg_from_master(path, ticker):
    """Return (bullish, bearish, call_walls, put_walls) for ticker from a master's
    Aggregate sheet, or None if the ticker/sheet/file is missing."""
    try:
        from openpyxl import load_workbook
    except Exception:
        return None
    if not path.exists():
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    if "Aggregate" not in wb.sheetnames:
        wb.close()
        return None
    ws = wb["Aggregate"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        wb.close()
        return None
    idx = {str(h): i for i, h in enumerate(header)}
    t_i = idx.get("ticker", 0)
    for row in rows:
        if row and str(row[t_i]).upper() == ticker:
            def num(col):
                i = idx.get(col)
                return parse_premium(row[i]) if i is not None and row[i] is not None else None
            bullish = num("bullish_premium") or 0.0
            bearish = num("bearish_premium") or 0.0
            walls = lambda pfx: [row[idx[c]] for c in (f"{pfx}_1", f"{pfx}_2", f"{pfx}_3")
                                 if c in idx and row[idx[c]] not in (None, "")]
            call_walls = [float(x) for x in walls("call_wall")]
            put_walls = [float(x) for x in walls("put_wall")]
            wb.close()
            return bullish, bearish, call_walls, put_walls
    wb.close()
    return None


def _ticker_in_master(path, ticker):
    agg = _agg_from_master(path, ticker)
    return agg is not None


def _agg_from_daycsv(directory, ticker):
    """Fallback: sum sentiment premium for ticker from today's live day-CSV."""
    day = datetime.now().strftime("%Y-%m-%d")
    path = directory / f"flow_{day}.csv"
    if not path.exists():
        return None
    bullish = bearish = 0.0
    hit = False
    try:
        with path.open(newline="") as fh:
            for r in csv.DictReader(fh):
                if str(r.get("ticker", "")).upper() != ticker:
                    continue
                prem = parse_premium(r.get("premium"), r.get("premiumTag")) or 0.0
                sent = str(r.get("sentiment", "")).lower()
                if sent in BULLISH_SENTS:
                    bullish += prem; hit = True
                elif sent in BEARISH_SENTS:
                    bearish += prem; hit = True
    except Exception:
        return None
    if not hit:
        return None
    return bullish, bearish, [], []


def conviction(ticker, directory):
    ticker = ticker.upper()
    directory = Path(directory)
    used = []

    # Fast path first.
    cached = _agg_from_cache(directory, ticker)
    if cached is not None:
        (bullish, bearish, call_walls, put_walls), in_unusual, in_knows = cached
        net = bullish - bearish
        total = bullish + bearish
        score = round(abs(net) / total, 4) if total > 0 else 0.0
        direction = "neutral" if total <= 0 else ("bullish" if net > 0 else "bearish" if net < 0 else "neutral")
        return {
            "ticker": ticker, "found": True,
            "bullish_premium": round(bullish, 2), "bearish_premium": round(bearish, 2),
            "net_premium": round(net, 2), "direction": direction, "score": score,
            "in_unusual": in_unusual, "in_knows": in_knows,
            "call_walls": call_walls, "put_walls": put_walls,
            "sources": [CACHE_NAME],
        }

    agg = _agg_from_master(directory / MASTERS["live"], ticker)
    if agg is not None:
        used.append(MASTERS["live"])
    else:
        agg = _agg_from_daycsv(directory, ticker)
        if agg is not None:
            used.append(f"flow_{datetime.now():%Y-%m-%d}.csv")

    if agg is None:
        return {"ticker": ticker, "found": False, "reason": "no OptionStrat data for ticker",
                "direction": "neutral", "score": 0.0}

    bullish, bearish, call_walls, put_walls = agg
    in_unusual = _ticker_in_master(directory / MASTERS["unusual"], ticker)
    in_knows = _ticker_in_master(directory / MASTERS["knows"], ticker)
    if in_unusual:
        used.append(MASTERS["unusual"])
    if in_knows:
        used.append(MASTERS["knows"])

    net = bullish - bearish
    total = bullish + bearish
    score = round(abs(net) / total, 4) if total > 0 else 0.0
    if total <= 0:
        direction = "neutral"
    elif net > 0:
        direction = "bullish"
    elif net < 0:
        direction = "bearish"
    else:
        direction = "neutral"

    return {
        "ticker": ticker,
        "found": True,
        "bullish_premium": round(bullish, 2),
        "bearish_premium": round(bearish, 2),
        "net_premium": round(net, 2),
        "direction": direction,
        "score": score,
        "in_unusual": in_unusual,
        "in_knows": in_knows,
        "call_walls": call_walls,
        "put_walls": put_walls,
        "sources": used,
    }


# ── DISCOVERY: rank the whole book, don't just answer about one ticker ──────

def _all_rows_from_master(path):
    """Every Aggregate row for a master as {ticker: (bullish, bearish)}."""
    out = {}
    try:
        from openpyxl import load_workbook
    except Exception:
        return out
    if not path.exists():
        return out
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return out
    if "Aggregate" not in wb.sheetnames:
        wb.close(); return out
    rows = ws_rows = wb["Aggregate"].iter_rows(values_only=True)
    header = next(ws_rows, None)
    if not header:
        wb.close(); return out
    idx = {str(h): i for i, h in enumerate(header)}
    t_i = idx.get("ticker", 0)
    b_i, r_i = idx.get("bullish_premium"), idx.get("bearish_premium")
    for row in ws_rows:
        if not row or row[t_i] in (None, ""):
            continue
        t = str(row[t_i]).upper()
        bull = parse_premium(row[b_i]) if b_i is not None else 0.0
        bear = parse_premium(row[r_i]) if r_i is not None else 0.0
        out[t] = (bull or 0.0, bear or 0.0)
    wb.close()
    return out


def _daycsv_all(directory):
    """Fallback discovery source: aggregate today's live day-CSV by ticker."""
    day = datetime.now().strftime("%Y-%m-%d")
    path = directory / f"flow_{day}.csv"
    out = {}
    if not path.exists():
        return out
    try:
        with path.open(newline="") as fh:
            for r in csv.DictReader(fh):
                t = str(r.get("ticker", "")).upper()
                if not t:
                    continue
                prem = parse_premium(r.get("premium"), r.get("premiumTag")) or 0.0
                sent = str(r.get("sentiment", "")).lower()
                bull, bear = out.get(t, (0.0, 0.0))
                if sent in BULLISH_SENTS:
                    bull += prem
                elif sent in BEARISH_SENTS:
                    bear += prem
                else:
                    continue
                out[t] = (bull, bear)
    except Exception:
        return {}
    return out


def discover(directory, top_n=10, min_premium=250_000.0, min_score=0.30,
             direction="bullish", boost_unusual=1.25, boost_knows=1.5):
    """Rank tickers by bullish conviction across all three feeds.

    rank = net_premium * skew_score * feed_boost
      net_premium  — bullish minus bearish (only positive/bullish names kept)
      skew_score   — |net| / total, so a one-sided book beats a near-tie
      feed_boost   — multiplier if the name also appears in the "Highly Unusual"
                     and/or "In The Know" books (your higher-signal presets)
    """
    directory = Path(directory)
    src = "masters"
    blob = _load_cache(directory)
    if blob:                                   # fast path
        live = {t: (e.get("bull", 0.0), e.get("bear", 0.0)) for t, e in blob["tickers"].items()}
        unusual = {t for t, e in blob["tickers"].items() if e.get("u")}
        knows = {t for t, e in blob["tickers"].items() if e.get("k")}
        src = CACHE_NAME
    else:
        live = _all_rows_from_master(directory / MASTERS["live"])
        if not live:
            live = _daycsv_all(directory)
            src = "day-csv"
        unusual = set(_all_rows_from_master(directory / MASTERS["unusual"]).keys())
        knows = set(_all_rows_from_master(directory / MASTERS["knows"]).keys())

    want_bull = direction in ("bullish", "both")
    want_bear = direction in ("bearish", "both")

    cands = []
    for t, (bull, bear) in live.items():
        net = bull - bear
        total = bull + bear
        if total <= 0:
            continue
        # Keep only the direction(s) the caller trades. A bullish book is a long
        # candidate; a bearish book is a short (put) candidate.
        if net > 0 and not want_bull:
            continue
        if net < 0 and not want_bear:
            continue
        if net == 0:
            continue
        score = abs(net) / total
        if abs(net) < min_premium or score < min_score:
            continue
        side = "long" if net > 0 else "short"
        boost = 1.0
        if t in unusual:
            boost *= boost_unusual
        if t in knows:
            boost *= boost_knows
        cands.append({
            "ticker": t,
            "side": side,
            "bullish_premium": round(bull, 2),
            "bearish_premium": round(bear, 2),
            "net_premium": round(net, 2),          # signed: +bullish, -bearish
            "abs_premium": round(abs(net), 2),     # magnitude, for ranking
            "score": round(score, 4),
            "in_unusual": t in unusual,
            "in_knows": t in knows,
            "rank": round(abs(net) * score * boost, 2),
        })
    cands.sort(key=lambda c: c["rank"], reverse=True)
    return {"source": "optionstrat", "via": src, "count": len(cands), "candidates": cands[:top_n]}


def main():
    argv = sys.argv[1:]
    # Discovery mode:  optionstrat_flow.py --discover <dir> [topN] [minPremium] [minScore]
    if argv and argv[0] == "--discover":
        directory = argv[1] if len(argv) > 1 else "."
        top_n = int(argv[2]) if len(argv) > 2 else 10
        min_prem = float(argv[3]) if len(argv) > 3 else 250_000.0
        min_score = float(argv[4]) if len(argv) > 4 else 0.30
        direction = argv[5] if len(argv) > 5 else "bullish"
        try:
            print(json.dumps(discover(directory, top_n, min_prem, min_score, direction)))
        except Exception as e:
            print(json.dumps({"source": "optionstrat", "count": 0, "candidates": [], "error": str(e)}))
        return

    if not argv:
        print(json.dumps({"found": False, "error": "usage: optionstrat_flow.py TICKER [dir] | --discover DIR [topN]"}))
        return
    ticker = argv[0]
    directory = argv[1] if len(argv) > 1 else "."
    try:
        print(json.dumps(conviction(ticker, directory)))
    except Exception as e:  # never crash the caller
        print(json.dumps({"ticker": ticker.upper(), "found": False, "error": str(e),
                          "direction": "neutral", "score": 0.0}))


if __name__ == "__main__":
    main()
