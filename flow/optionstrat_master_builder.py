"""
OptionStrat master-file builder — separate from the scraper.

WHY a separate script: optionstrat_flow_scraper.py's only job is reliably capturing
live rows into per-day CSVs. This script's job is turning those day-files into
persistent, deduped, auto-pruned master workbooks. Keeping them apart means a bug in
one never risks the other, and each stays simple enough to actually debug.

WHAT IT DOES
Reads today's day-file for each feed (live / unusual / knows) and updates a matching
master .xlsx (flow_master.xlsx, flow_unusual_master.xlsx, flow_knows_master.xlsx).
Each master workbook has two sheets:

  "Active"     - every non-expired row ever captured for that feed, deduped.
                 New rows get appended; rows whose option has expired are DELETED
                 outright (not archived) each time this script runs.
  "Aggregate"  - one row per ticker, computed fresh from Active each run:
                 bullish_premium / bearish_premium / neutral_premium (sentiment-
                 based: bullish+very-bullish vs bearish+very-bearish; neutral kept
                 separate, not folded into either side), plus the top 3 call-side
                 and put-side strikes by premium concentration ("walls").

STRIKE / OPTION-TYPE PARSING
The scraper doesn't capture a dedicated strike field - it's embedded in the
free-text `strategy` column, e.g.:
  "Buy 285 Put To Open"        -> strike 285, PUT
  "Sell 520 Call To Open"      -> strike 520, CALL
  "Sell 140/200 Strangle"      -> two legs: PUT @140, CALL @200 (split in half)
  "Sell 140/200 Combo"         -> same treatment as Strangle
  "Buy 100/110/120 Butterfly"  -> 3 legs, alternating put/call assignment (best
                                   effort - butterflies/condors aren't cleanly
                                   directional, flagged via `multi_leg=True`)
Rows where no strike can be parsed are kept in Active (for the record) but excluded
from the Aggregate wall/premium math.

PREMIUM PARSING: "$1.16m" / "$950k" -> numeric USD.

EXPIRATION PARSING: handles both "18-Dec" (day-Mon, assumes current/next year) and
"Dec 15 '28" (Mon day 'YY) formats seen in the sample data.

Setup:  pip install openpyxl
Run:    python optionstrat_master_builder.py
        (run after/alongside the scraper - reads whatever day-files exist for today)
"""

import csv
import re
from datetime import datetime, date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

FEEDS = [
    {"label": "live",    "day_prefix": "flow",         "master": "flow_master.xlsx"},
    {"label": "unusual",  "day_prefix": "flow_unusual", "master": "flow_unusual_master.xlsx"},
    {"label": "knows",    "day_prefix": "flow_knows",   "master": "flow_knows_master.xlsx"},
]

DAY_FIELDS = ["captured", "ticker", "premium", "premiumTag", "strategy", "badge",
              "expiration", "time", "sentiment"]

# Physical column order in the Active sheet = display columns first (in the exact
# order requested), then the hidden/internal-logic-only columns after. Hiding is
# applied via column_dimensions[...].hidden, not by removing the data - de-dupe
# (row_key uses raw `expiration` + `time`), sort (`premium_usd`), and aggregate
# (`strike`, `option_type`, `leg_count`) all still need these fields present.
ACTIVE_DISPLAY_ORDER = ["ticker", "strategy", "badge_tag", "premium", "expiration_date",
                         "sentiment", "leg_count", "time"]
ACTIVE_HIDDEN_ORDER = ["premiumTag", "badge", "expiration", "strike", "option_type", "premium_usd"]
ACTIVE_FIELDS = ACTIVE_DISPLAY_ORDER + ACTIVE_HIDDEN_ORDER

BULLISH_SENTS = {"bullish", "very-bullish"}
BEARISH_SENTS = {"bearish", "very-bearish"}

KEEP_DAYS = 1  # delete day-files older than this many days once merged into the master

TOP_N_OVERALL = 10  # how many single largest-premium rows to spotlight in the Top 10 sheet

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

# How close bullish_premium and bearish_premium must be (as a fraction of the
# larger of the two) to count as a "near-tie" and get left uncolored, instead of
# a razor-thin skew getting the same green/red as a completely lopsided ticker.
SKEW_NEUTRAL_BAND = 0.10  # 10%

MONTHS = {m.lower(): i for i, m in enumerate(
    ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"], start=1)}


# ── parsing helpers ──────────────────────────────────────────────────────────

def parse_premium(premium_str, tag):
    """'$1.16m' -> 1160000.0. Tag ('K'/'M'/etc, if the site splits it out) is used
    only if the numeric string itself has no suffix."""
    if not premium_str:
        return None
    s = premium_str.strip().lstrip("$").lower()
    mult = 1.0
    if s.endswith("m"):
        mult, s = 1_000_000.0, s[:-1]
    elif s.endswith("k"):
        mult, s = 1_000.0, s[:-1]
    elif s.endswith("b"):
        mult, s = 1_000_000_000.0, s[:-1]
    elif tag:
        t = tag.strip().lower()
        if t == "m":
            mult = 1_000_000.0
        elif t == "k":
            mult = 1_000.0
        elif t == "b":
            mult = 1_000_000_000.0
    try:
        return float(s) * mult
    except ValueError:
        return None


def parse_expiration(exp_str, today):
    """Handles the real formats seen in OptionStrat's export:
      "Jul 6"        - month-name + day, no year  (most common)
      "Jul 10 - Aug 7" - a date range (calendar spread); we use the LATER date,
                         since that's when the position is fully closed out
      "Dec 15 '28"   - month day 'YY
      "18-Dec"       - day-Mon (older/alternate format, kept for safety)
    All "no year given" cases assume current year, rolling to next year if that
    date is more than ~60 days in the past (handles December -> January turnover).
    """
    if not exp_str:
        return None
    exp_str = exp_str.strip()

    # "Jul 10 - Aug 7" -> take the later side of the range
    if "-" in exp_str and re.search(r"[A-Za-z]{3}\s+\d{1,2}\s*-\s*[A-Za-z]{3}\s+\d{1,2}", exp_str):
        parts = [p.strip() for p in exp_str.split("-", 1)]
        d2 = parse_expiration(parts[1], today) if len(parts) == 2 else None
        d1 = parse_expiration(parts[0], today) if len(parts) == 2 else None
        if d1 and d2:
            return max(d1, d2)
        return d2 or d1

    # "Dec 15 '28"
    m = re.match(r"([A-Za-z]{3})\s+(\d{1,2})\s+'(\d{2})", exp_str)
    if m:
        mon, day, yy = m.groups()
        mon_num = MONTHS.get(mon.lower())
        if mon_num:
            year = 2000 + int(yy)
            try:
                return date(year, mon_num, int(day))
            except ValueError:
                return None

    # "Jul 6" / "Sep 18" -> month-name + day, no year
    m = re.match(r"([A-Za-z]{3})\w*\s+(\d{1,2})$", exp_str)
    if m:
        mon, day = m.groups()
        mon_num = MONTHS.get(mon.lower()[:3])
        if mon_num:
            year = today.year
            try:
                d = date(year, mon_num, int(day))
            except ValueError:
                return None
            if (today - d).days > 60:
                try:
                    d = date(year + 1, mon_num, int(day))
                except ValueError:
                    pass
            return d

    # "18-Dec" or "18-Dec '25" -> day-Mon, optional 'YY year suffix
    m = re.match(r"(\d{1,2})-([A-Za-z]{3})(?:\s*'(\d{2}))?$", exp_str)
    if m:
        day, mon, yy = m.groups()
        mon_num = MONTHS.get(mon.lower())
        if mon_num:
            if yy:
                year = 2000 + int(yy)
            else:
                # no year given; assume current year, roll to next year if that
                # date has already passed by more than ~60 days (turn-of-year case)
                year = today.year
                try:
                    d = date(year, mon_num, int(day))
                except ValueError:
                    return None
                if (today - d).days > 60:
                    year += 1
            try:
                return date(year, mon_num, int(day))
            except ValueError:
                return None
    return None


_STRAT_RE = re.compile(
    r"(?P<action>Buy|Sell)\s+(?P<strikes>[\d./]+)\s+(?P<kind>Puts?|Calls?|Strangle|Combo|Straddle|Butterfly|Condor)",
    re.IGNORECASE,
)
# "155/165 Bear Put Spread", "0.5/1 Bull Call Spread To Open",
# "155/160 Diagonal Call Spread", "230/255 Reverse Diagonal Call Spread",
# "785/820 Call Ratio Spread", "85 Call Ratio Spread" (single-strike ratio spread -
# same strike bought/sold in different quantities) - no leading Buy/Sell; Put/Call
# before "Spread" tells us the type. The directional/ratio modifier word doesn't
# change which side (put vs call) the strikes belong to.
_SPREAD_RE = re.compile(
    r"(?P<strikes>[\d./]+)\s+(?:Bull|Bear|Diagonal|Reverse\s+Diagonal)?\s*(?P<kind>Put|Call)\s+(?:Ratio\s+)?Spread",
    re.IGNORECASE,
)
# "21 Reverse Calendar Put Spread", "730 Calendar Call Spread" - single strike,
# both legs at the same strike but different expiries. For wall purposes this is
# one strike concentration point, so treat it as a single leg (not doubled).
_CALENDAR_RE = re.compile(
    r"(?P<strikes>[\d./]+)\s+(?:Reverse\s+)?Calendar\s+(?P<kind>Put|Call)\s+Spread",
    re.IGNORECASE,
)
# "315 Straddle", "175/202.5 Strangle", "165/195 Strangle" - no leading Buy/Sell.
_BARE_STRANGLE_RE = re.compile(
    r"^(?P<strikes>[\d./]+)\s+(?P<kind>Strangle|Straddle)\b",
    re.IGNORECASE,
)
# "Sell 104/117 Guts", "Sell 115/155 Guts" - like a Strangle/Combo but ITM legs
# (call strike BELOW put strike, i.e. they cross through the money) instead of
# both OTM. Put/call assignment is the same idea: lower strike -> call leg,
# higher strike -> put leg (opposite of Strangle/Combo, since Guts crosses ITM).
_GUTS_RE = re.compile(
    r"(?:Buy|Sell)?\s*(?P<strikes>[\d./]+)\s+Guts",
    re.IGNORECASE,
)
# "15/25/40/60 Iron Condor To Open", "1600/1780/1920/2020 Iron Condor",
# "420/425/430/435 Inverse Iron Condor" - 4 strikes: put spread (two lowest
# strikes) + call spread (two highest strikes). "Inverse"/"Reverse" flips long
# vs short but the strike-to-side assignment (lowest two = puts, highest two =
# calls) is the same either way for wall purposes.
_IRON_CONDOR_RE = re.compile(
    r"(?:Buy|Sell)?\s*(?P<strikes>[\d./]+)\s+(?:Inverse\s+|Reverse\s+)?Iron\s+Condor",
    re.IGNORECASE,
)
# INTENTIONALLY NOT PARSED (kept in Active, excluded from Aggregate wall/premium math):
#   "Synthetic Future"  - net delta ~1 position (long call + short put at one strike,
#                          or the reverse) - it's a stock-replacement play, not a
#                          directional call/put bet, so it doesn't belong in either wall.
#   "Call/Put Ratio Spread" - unequal leg counts (e.g. 1x2); rare, low priority.
# If any of these become frequent in your data, tell me and I'll add real handling.


def parse_strategy(strategy):
    """Returns list of (strike, option_type) legs, and a multi_leg flag.

    Handles the shapes actually seen in OptionStrat exports:
      "Buy 285 Put To Open"          -> [(285, PUT)]
      "Sell 520 Call To Open"        -> [(520, CALL)]
      "Sell 726/727 Puts"            -> [(726, PUT), (727, PUT)]   (plural = same side,
                                          NOT a strangle - both legs are puts)
      "Sell 90/91 Calls"             -> [(90, CALL), (91, CALL)]
      "Sell 140/200 Strangle"/"Combo"-> [(140, PUT), (200, CALL)] (split: lower strike
                                          = put leg, higher = call leg)
      "155/165 Bear Put Spread"      -> [(155, PUT), (165, PUT)]  (no Buy/Sell prefix)
      "0.5/1 Bull Call Spread"       -> [(0.5, CALL), (1, CALL)]
    Straddles/butterflies/condors handled as before (best-effort).
    Rows that don't match any known shape return ([], False) and are kept in Active
    for the record but excluded from the Aggregate premium/wall math.
    """
    if not strategy:
        return [], False

    # "<strikes> Bull/Bear/Diagonal Put/Call Spread" - no Buy/Sell prefix
    m = _SPREAD_RE.search(strategy)
    if m:
        opt_type = m.group("kind").upper()
        strikes = _parse_strike_list(m.group("strikes"))
        if strikes:
            return [(s, opt_type) for s in strikes], len(strikes) > 1

    # "<strike> (Reverse) Calendar Put/Call Spread" - same strike, different expiries;
    # one strike concentration point, so a single leg (not doubled for wall purposes).
    m = _CALENDAR_RE.search(strategy)
    if m:
        opt_type = m.group("kind").upper()
        strikes = _parse_strike_list(m.group("strikes"))
        if strikes:
            return [(strikes[0], opt_type)], False

    # "315 Straddle", "175/202.5 Strangle" - no leading Buy/Sell
    m = _BARE_STRANGLE_RE.match(strategy)
    if m:
        kind = m.group("kind").lower()
        strikes = _parse_strike_list(m.group("strikes"))
        if kind == "straddle" and strikes:
            return [(strikes[0], "PUT"), (strikes[0], "CALL")], True
        if kind == "strangle" and len(strikes) >= 2:
            lo, hi = min(strikes), max(strikes)
            return [(lo, "PUT"), (hi, "CALL")], True

    # "Sell 104/117 Guts" - ITM strangle: call strike is the LOWER one (below the
    # put strike), since the legs cross through the money. Opposite assignment
    # from Strangle/Combo.
    m = _GUTS_RE.search(strategy)
    if m:
        strikes = _parse_strike_list(m.group("strikes"))
        if len(strikes) >= 2:
            lo, hi = min(strikes), max(strikes)
            return [(lo, "CALL"), (hi, "PUT")], True

    # "15/25/40/60 Iron Condor" - 4 strikes sorted ascending: two lowest = put
    # wing, two highest = call wing.
    m = _IRON_CONDOR_RE.search(strategy)
    if m:
        strikes = sorted(_parse_strike_list(m.group("strikes")))
        if len(strikes) == 4:
            put_lo, put_hi, call_lo, call_hi = strikes
            return [(put_lo, "PUT"), (put_hi, "PUT"), (call_lo, "CALL"), (call_hi, "CALL")], True
        elif len(strikes) >= 2:
            # fewer than 4 strikes given (unusual) - fall back to a simple
            # lower-half-put / upper-half-call split so it's not lost entirely
            mid = len(strikes) // 2
            legs = [(s, "PUT") for s in strikes[:mid]] + [(s, "CALL") for s in strikes[mid:]]
            return legs, True

    m = _STRAT_RE.search(strategy)
    if not m:
        return [], False

    kind = m.group("kind").lower().rstrip("s")  # normalize "Puts"/"Calls" -> "put"/"call"
    is_plural = m.group("kind").lower().endswith("s") and kind in ("put", "call")
    strikes = _parse_strike_list(m.group("strikes"))
    if not strikes:
        return [], False

    if kind in ("put", "call"):
        if len(strikes) > 1 and is_plural:
            # "Sell 726/727 Puts" - same side, multiple strikes (e.g. a put spread
            # written without the word "Spread"), not one put + one call.
            return [(s, kind.upper()) for s in strikes], True
        return [(strikes[0], kind.upper())], False

    if kind == "straddle":
        return [(strikes[0], "PUT"), (strikes[0], "CALL")], True

    if kind in ("strangle", "combo") and len(strikes) >= 2:
        lo, hi = min(strikes), max(strikes)
        return [(lo, "PUT"), (hi, "CALL")], True

    if kind in ("butterfly", "condor") and len(strikes) >= 2:
        # best-effort: alternate put/call assignment across sorted strikes
        strikes_sorted = sorted(strikes)
        legs = [(k, "PUT" if i % 2 == 0 else "CALL") for i, k in enumerate(strikes_sorted)]
        return legs, True

    return [], False


def combine_badge_tag(badge, tag):
    """'Sweep' + 'aa' -> 'Sweep-aa' (case preserved as-is). Badge-only -> 'Sweep'.
    Tag-only -> 'aa'. Neither -> ''."""
    badge = (badge or "").strip()
    tag = (tag or "").strip()
    if badge and tag:
        return f"{badge}-{tag}"
    return badge or tag or ""


def _parse_strike_list(strikes_raw):
    strikes = []
    for tok in strikes_raw.split("/"):
        try:
            strikes.append(float(tok))
        except ValueError:
            pass
    return strikes


# ── IO helpers ────────────────────────────────────────────────────────────────

def today_str():
    return datetime.now().strftime("%Y-%m-%d")


def find_day_files(prefix):
    """Find all day-files matching this feed's prefix, e.g. flow_2026-07-02.csv,
    flow_unusual_2026-07-02.csv. Returns list of (day_str, path), oldest first.
    Guards against the 'live' feed's plain 'flow' prefix accidentally matching
    'flow_unusual_*'/'flow_knows_*' files via a naive glob."""
    out = []
    for p in Path(".").glob(f"{prefix}_*.csv"):
        m = re.match(rf"^{re.escape(prefix)}_(\d{{4}}-\d{{2}}-\d{{2}})\.csv$", p.name)
        if m:
            out.append((m.group(1), p))
    return sorted(out, key=lambda t: t[0])


def read_day_file(path):
    rows = []
    if path.exists():
        with path.open(newline="") as f:
            for r in csv.DictReader(f):
                rows.append(r)
    return rows


def row_key(r):
    return (r["ticker"], r["strategy"], r["expiration"], r["time"], r["premium"])


def load_active_sheet(wb):
    if "Active" not in wb.sheetnames:
        ws = wb.create_sheet("Active")
        ws.append(ACTIVE_FIELDS)
        return ws, set()
    ws = wb["Active"]
    seen = set()
    header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    if header != ACTIVE_FIELDS:
        # header drift (first run / schema change) - rebuild cleanly, keep no stale cols
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.remove(ws)
        ws = wb.create_sheet("Active")
        ws.append(ACTIVE_FIELDS)
        for row in rows:
            d = dict(zip(header, row))
            seen.add(row_key({k: (d.get(k) or "") for k in DAY_FIELDS}))
            ws.append([d.get(f) for f in ACTIVE_FIELDS])
    else:
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(ACTIVE_FIELDS, row))
            seen.add(row_key({k: (d.get(k) or "") for k in DAY_FIELDS}))
    return ws, seen


def prune_expired(ws, today):
    """Delete rows whose parsed expiration_date has passed.
    Returns (removed_count, removed_keys) - removed_keys lets the caller avoid
    re-adding the same expired row again this run if it's still in today's day-file.

    Rebuilds the sheet in a single pass (read all rows once, delete_rows(2, max),
    re-append survivors) rather than calling delete_rows() once per expired row.
    openpyxl's delete_rows() is O(n) per call (shifts every row below up by one),
    so deleting one-at-a-time in a loop is O(n^2) overall and is what caused this
    script to effectively hang on real data volumes - see prune_neutral_tickers()
    for the same fix applied there."""
    exp_idx = ACTIVE_FIELDS.index("expiration_date")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return 0, set()

    survivors = []
    removed_keys = set()
    for row in rows:
        val = row[exp_idx]
        is_expired = False
        if val:
            try:
                d = val if isinstance(val, date) else datetime.strptime(str(val), "%Y-%m-%d").date()
                is_expired = d < today
            except ValueError:
                pass
        if is_expired:
            row_dict = dict(zip(ACTIVE_FIELDS, row))
            removed_keys.add(row_key({k: (row_dict.get(k) or "") for k in DAY_FIELDS}))
        else:
            survivors.append(row)

    removed = len(rows) - len(survivors)
    if removed:
        ws.delete_rows(2, ws.max_row)
        for row in survivors:
            ws.append(row)

    return removed, removed_keys


def sort_active_sheet(ws):
    """Reorder Active's data rows by ticker (A-Z), then premium_usd (largest first)
    within each ticker. Rewrites the sheet body in place after all appends/prunes
    for this run are done."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return
    ticker_idx = ACTIVE_FIELDS.index("ticker")
    prem_idx = ACTIVE_FIELDS.index("premium_usd")

    def sort_key(row):
        ticker = row[ticker_idx] or ""
        try:
            prem = float(row[prem_idx]) if row[prem_idx] not in (None, "") else 0.0
        except (TypeError, ValueError):
            prem = 0.0
        return (ticker, -prem)  # ticker ascending, premium descending

    rows.sort(key=sort_key)
    ws.delete_rows(2, ws.max_row)
    for row in rows:
        ws.append(row)


def apply_column_display(ws):
    """Set sensible widths for the visible Active columns and hide the
    internal-logic-only ones (they're still real data, just not shown)."""
    for f in ACTIVE_DISPLAY_ORDER:
        col_letter = get_column_letter(ACTIVE_FIELDS.index(f) + 1)
        ws.column_dimensions[col_letter].width = 14 if f not in ("strategy",) else 30
    for f in ACTIVE_HIDDEN_ORDER:
        col_letter = get_column_letter(ACTIVE_FIELDS.index(f) + 1)
        ws.column_dimensions[col_letter].hidden = True


def cleanup_old_day_files(feed, day_files, keep_days, today_date):
    """Delete day-files older than `keep_days` days, but ONLY after this run has
    already merged them into the master (they're in `day_files`, which means
    read_day_file() was just called on them this run). Never deletes the file for
    today or for a day still within the keep window."""
    deleted = []
    for day_str, path in day_files:
        try:
            file_date = datetime.strptime(day_str, "%Y-%m-%d").date()
        except ValueError:
            continue
        age_days = (today_date - file_date).days
        if age_days > keep_days:
            try:
                path.unlink()
                deleted.append(path.name)
            except OSError as e:
                print(f"  [warn] could not delete {path}: {e}")
    return deleted


SUPPRESSED_SHEET = "_PrunedNeutralKeys"  # hidden bookkeeping sheet, not meant for viewing


def load_pruned_neutral_keys(wb):
    """Read the persisted set of row-keys that were removed by neutral-ticker
    pruning as of the end of the last run. Needed because once a row is deleted
    from Active, `seen` (rebuilt fresh from Active each run) no longer contains
    its key, so without this the exact same day-file row would be re-added and
    then immediately re-pruned every single run for as long as its ticker stays
    neutral and the day-file still exists. Mirrors how prune_expired() already
    returns removed_keys for the same reason. Returns a set of row_key tuples
    (empty set if the sheet doesn't exist yet, e.g. first run)."""
    if SUPPRESSED_SHEET not in wb.sheetnames:
        return set()
    ws = wb[SUPPRESSED_SHEET]
    return {tuple(row) for row in ws.iter_rows(min_row=1, values_only=True) if row and row[0]}


def save_pruned_neutral_keys(wb, keys):
    """Overwrite the persisted pruned-neutral-key list with the current set."""
    if SUPPRESSED_SHEET in wb.sheetnames:
        wb.remove(wb[SUPPRESSED_SHEET])
    ws = wb.create_sheet(SUPPRESSED_SHEET)
    ws.sheet_state = "hidden"
    for k in keys:
        ws.append(list(k))


def compute_neutral_tickers(rows_as_dicts, neutral_band=SKEW_NEUTRAL_BAND):
    """Given a list of row-dicts (keyed by ACTIVE_FIELDS), compute which tickers
    are a bullish/bearish near-tie (same math as build_aggregate/prune_neutral_tickers).
    Returns a set of ticker strings. Pure function, no sheet I/O - lets the caller
    reuse the same computation both for "which existing Active rows to drop" and
    for "which day-file rows to pre-skip this run" without parsing anything twice."""
    totals = {}
    for d in rows_as_dicts:
        ticker = d.get("ticker")
        if not ticker:
            continue
        sentiment = (d.get("sentiment") or "").strip().lower()
        try:
            prem = float(d.get("premium_usd")) if d.get("premium_usd") not in (None, "") else 0.0
        except (TypeError, ValueError):
            prem = 0.0
        t = totals.setdefault(ticker, {"bullish": 0.0, "bearish": 0.0})
        if sentiment in BULLISH_SENTS:
            t["bullish"] += prem
        elif sentiment in BEARISH_SENTS:
            t["bearish"] += prem

    neutral_tickers = set()
    for ticker, t in totals.items():
        larger = max(t["bullish"], t["bearish"])
        if larger == 0:
            continue
        diff_ratio = abs(t["bullish"] - t["bearish"]) / larger
        if diff_ratio < neutral_band:
            neutral_tickers.add(ticker)
    return neutral_tickers


def prune_neutral_tickers(ws, neutral_band=SKEW_NEUTRAL_BAND):
    """Delete ALL Active rows for any ticker whose bullish_premium and
    bearish_premium are a near-tie (within `neutral_band`, same threshold used
    for the Aggregate skew highlighting) - i.e. a ticker with no real net
    directional lean, "just stuck there." Computed the same way build_aggregate()
    computes bullish/bearish premium, so the two stay consistent.

    Rebuilds the sheet in a single pass (read all rows once, delete_rows(2, max),
    re-append survivors) rather than calling delete_rows() once per deleted row -
    openpyxl's delete_rows() is O(n) per call (it shifts every row below up by
    one), so deleting one row at a time on a large sheet is effectively O(n^2)
    and can turn a few-second run into something that never finishes on real
    data volumes. This mirrors the same one-pass approach sort_active_sheet()
    already uses.

    Returns (removed_row_count, removed_ticker_set, removed_row_keys)."""
    ticker_idx = ACTIVE_FIELDS.index("ticker")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return 0, set(), set()

    rows_as_dicts = [dict(zip(ACTIVE_FIELDS, row)) for row in rows]
    neutral_tickers = compute_neutral_tickers(rows_as_dicts, neutral_band)

    if not neutral_tickers:
        return 0, set(), set()

    survivors = []
    removed_row_keys = set()
    for row, d in zip(rows, rows_as_dicts):
        if row[ticker_idx] in neutral_tickers:
            removed_row_keys.add(row_key({k: (d.get(k) or "") for k in DAY_FIELDS}))
        else:
            survivors.append(row)
    removed_rows = len(rows) - len(survivors)

    ws.delete_rows(2, ws.max_row)
    for row in survivors:
        ws.append(row)

    return removed_rows, neutral_tickers, removed_row_keys


def apply_skew_highlighting(ws):
    """Green fill on rows where bullish_premium clearly exceeds bearish_premium,
    red where bearish clearly exceeds bullish. Near-ties (within SKEW_NEUTRAL_BAND
    of each other) are left unhighlighted so a razor-thin skew isn't colored as
    strongly as a lopsided one."""
    bull_col = 2  # 'bullish_premium' is always column B in Aggregate's fixed header
    bear_col = 3  # 'bearish_premium' is always column C
    for r in range(2, ws.max_row + 1):
        bull = ws.cell(row=r, column=bull_col).value or 0
        bear = ws.cell(row=r, column=bear_col).value or 0
        larger = max(bull, bear)
        if larger == 0:
            continue
        diff_ratio = abs(bull - bear) / larger
        if diff_ratio < SKEW_NEUTRAL_BAND:
            continue  # near-tie, leave uncolored
        fill = GREEN_FILL if bull > bear else RED_FILL
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill


def build_top10_sheet(wb, ws_active, max_tickers=TOP_N_OVERALL, max_per_ticker=2):
    """New 'Top 10 Overall' sheet: the top `max_tickers` DISTINCT TICKERS ranked by
    their largest single premium_usd, each contributing UP TO `max_per_ticker` rows
    (its 2 biggest prints, if it has that many). Total row count can therefore
    exceed max_tickers (e.g. 10 tickers x up to 2 rows each = up to 20 rows) - the
    cap is on ticker COUNT, not row count. Walks the full premium-ranked list;
    the first time a new ticker is seen it "claims" one of the max_tickers slots,
    then can contribute up to max_per_ticker rows total before being excluded from
    further additions. Also highlights those same rows gold directly in Active so
    they're visible in place without needing to jump sheets."""
    if "Top 10 Overall" in wb.sheetnames:
        wb.remove(wb["Top 10 Overall"])
    ws_top = wb.create_sheet("Top 10 Overall")
    ws_top.append(ACTIVE_DISPLAY_ORDER)

    prem_col = ACTIVE_FIELDS.index("premium_usd") + 1
    ticker_col = ACTIVE_FIELDS.index("ticker") + 1
    display_cols = [ACTIVE_FIELDS.index(f) + 1 for f in ACTIVE_DISPLAY_ORDER]

    rows_with_prem = []
    for r in range(2, ws_active.max_row + 1):
        prem_val = ws_active.cell(row=r, column=prem_col).value
        try:
            prem = float(prem_val) if prem_val not in (None, "") else 0.0
        except (TypeError, ValueError):
            prem = 0.0
        ticker = ws_active.cell(row=r, column=ticker_col).value
        rows_with_prem.append((r, prem, ticker))

    rows_with_prem.sort(key=lambda t: t[1], reverse=True)

    top_rows = []
    per_ticker_count = {}
    for r, prem, ticker in rows_with_prem:
        count = per_ticker_count.get(ticker, 0)
        if count == 0 and len(per_ticker_count) >= max_tickers:
            continue  # already have max_tickers distinct tickers claimed; skip new ones
        if count >= max_per_ticker:
            continue  # this ticker already has its max rows
        top_rows.append(r)
        per_ticker_count[ticker] = count + 1

    for r in top_rows:
        values = [ws_active.cell(row=r, column=c).value for c in display_cols]
        ws_top.append(values)
        for c in range(1, ws_active.max_column + 1):
            ws_active.cell(row=r, column=c).fill = GOLD_FILL

    for i, _ in enumerate(ACTIVE_DISPLAY_ORDER, start=1):
        ws_top.column_dimensions[get_column_letter(i)].width = 16


def build_aggregate(wb, active_rows):
    """Recompute the Aggregate sheet from scratch from current Active rows.
    Wall columns show ONLY the strike (call_wall_1 = highest-premium call strike,
    call_wall_2 = second, call_wall_3 = third; same idea for puts) - no premium
    sub-columns. Rows are highlighted green if bullish_premium > bearish_premium
    (net call-side skew) or red if bearish_premium > bullish_premium, with a
    neutral (unhighlighted) band for near-ties so a razor-thin skew doesn't get
    colored as strongly as a lopsided one."""
    if "Aggregate" in wb.sheetnames:
        wb.remove(wb["Aggregate"])
    ws = wb.create_sheet("Aggregate")
    header = ["ticker", "bullish_premium", "bearish_premium",
              "call_wall_1", "call_wall_2", "call_wall_3",
              "put_wall_1", "put_wall_2", "put_wall_3"]
    ws.append(header)

    by_ticker = {}
    for r in active_rows:
        ticker = r.get("ticker")
        if not ticker:
            continue
        agg = by_ticker.setdefault(ticker, {
            "bullish": 0.0, "bearish": 0.0,
            "call_strikes": {}, "put_strikes": {},
        })
        sentiment = (r.get("sentiment") or "").strip().lower()
        premium_usd = r.get("premium_usd")
        try:
            premium_usd = float(premium_usd) if premium_usd not in (None, "") else 0.0
        except (TypeError, ValueError):
            premium_usd = 0.0

        if sentiment in BULLISH_SENTS:
            agg["bullish"] += premium_usd
        elif sentiment in BEARISH_SENTS:
            agg["bearish"] += premium_usd
        # neutral sentiment: not tracked in Aggregate (dropped per user request)

        strike = r.get("strike")
        opt_type = r.get("option_type")
        if strike not in (None, "") and opt_type in ("CALL", "PUT"):
            try:
                strike_f = float(strike)
            except (TypeError, ValueError):
                strike_f = None
            if strike_f is not None:
                bucket = agg["call_strikes"] if opt_type == "CALL" else agg["put_strikes"]
                try:
                    leg_count = int(r.get("leg_count") or 1)
                except (TypeError, ValueError):
                    leg_count = 1
                contrib = premium_usd / leg_count if leg_count > 1 else premium_usd
                bucket[strike_f] = bucket.get(strike_f, 0.0) + contrib

    for ticker, agg in sorted(by_ticker.items()):
        top_calls = [s for s, _ in sorted(agg["call_strikes"].items(), key=lambda kv: kv[1], reverse=True)[:3]]
        top_puts = [s for s, _ in sorted(agg["put_strikes"].items(), key=lambda kv: kv[1], reverse=True)[:3]]
        top_calls += [None] * (3 - len(top_calls))
        top_puts += [None] * (3 - len(top_puts))
        row = [ticker, round(agg["bullish"], 2), round(agg["bearish"], 2)] + top_calls + top_puts
        ws.append(row)

    for i, _ in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    apply_skew_highlighting(ws)


# ── main ──────────────────────────────────────────────────────────────────────

def process_feed(feed, today_date):
    day_files = find_day_files(feed["day_prefix"])
    if not day_files:
        print(f"[{feed['label']}] no day-files found for prefix '{feed['day_prefix']}_*.csv' - skipping")
        return

    master_path = Path(feed["master"])
    wb = load_workbook(master_path) if master_path.exists() else Workbook()
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        wb.remove(wb["Sheet"])

    ws_active, seen = load_active_sheet(wb)
    removed, _ = prune_expired(ws_active, today_date)

    # Seed `seen` with row-keys that were removed by neutral-ticker pruning as of
    # the end of the LAST run (persisted in a hidden sheet - see
    # load_pruned_neutral_keys). Without this, those exact day-file rows would be
    # re-added and then immediately re-pruned again every single run for as long
    # as the ticker stays neutral and the day-file still exists, since `seen` is
    # otherwise rebuilt fresh from Active each run and no longer contains keys for
    # rows that were already deleted.
    previously_pruned_keys = load_pruned_neutral_keys(wb)
    seen |= previously_pruned_keys

    added = 0
    skipped_expired = 0
    skipped_neutral = 0
    unparsed_strategies = {}
    for day_str, day_path in day_files:
        for r in read_day_file(day_path):
            k = row_key(r)
            if k in seen:
                if k in previously_pruned_keys:
                    skipped_neutral += 1
                continue
            seen.add(k)

            legs, multi_leg = parse_strategy(r.get("strategy", ""))
            premium_usd = parse_premium(r.get("premium"), r.get("premiumTag"))
            exp_date = parse_expiration(r.get("expiration"), today_date)
            exp_date_str = exp_date.isoformat() if exp_date else None
            if r.get("expiration") and exp_date is None:
                print(f"  [warn] could not parse expiration '{r.get('expiration')}' "
                      f"for {r.get('ticker')} (strategy: '{r.get('strategy')}') - row kept but won't auto-expire")
            if not legs and r.get("strategy"):
                unparsed_strategies[r.get("strategy")] = unparsed_strategies.get(r.get("strategy"), 0) + 1

            # Skip rows that are already expired as of today - no point adding them
            # just to prune them again next run (this is what caused churn before).
            if exp_date is not None and exp_date < today_date:
                skipped_expired += 1
                continue

            if not legs:
                row_dict = dict(r)
                row_dict.update(strike=None, option_type=None, premium_usd=premium_usd,
                                 expiration_date=exp_date_str, leg_count=1,
                                 badge_tag=combine_badge_tag(r.get("badge"), r.get("premiumTag")))
                ws_active.append([row_dict.get(f) for f in ACTIVE_FIELDS])
                added += 1
                continue

            leg_count = len(legs)
            badge_tag = combine_badge_tag(r.get("badge"), r.get("premiumTag"))
            for strike, opt_type in legs:
                row_dict = dict(r)
                row_dict.update(strike=strike, option_type=opt_type, premium_usd=premium_usd,
                                 expiration_date=exp_date_str, leg_count=leg_count, badge_tag=badge_tag)
                ws_active.append([row_dict.get(f) for f in ACTIVE_FIELDS])
                added += 1

    # Remove tickers that are a net-zero wash (bullish premium ~= bearish premium,
    # within the same 10% neutral band used for Aggregate's skew coloring) - these
    # have no real directional lean and would otherwise just sit there forever.
    # Removing from Active here means they're automatically excluded from
    # Aggregate/Top10 too, since those are built from Active afterward.
    #
    # The persisted "pruned neutral keys" list must be the UNION of what was
    # already known-pruned (previously_pruned_keys, which we skipped re-adding
    # this run and therefore never had a chance to prune again) plus whatever
    # got freshly pruned just now - not just this run's fresh prunes alone.
    # Otherwise a run where nothing NEW needs pruning (because it was already
    # correctly pre-skipped) would overwrite the list with an empty set and
    # undo the pre-skipping on the very next run.
    neutral_removed, neutral_tickers, pruned_neutral_keys = prune_neutral_tickers(ws_active)
    all_pruned_keys = previously_pruned_keys | pruned_neutral_keys
    save_pruned_neutral_keys(wb, all_pruned_keys)

    # NOTE: no manual fill-clearing needed here - sort_active_sheet() below does a
    # full delete_rows()+append() rewrite of the sheet body, and openpyxl's
    # delete_rows() already discards cell styling (verified: a deleted row's fill
    # does not carry over when the row index is reused by a later append). An
    # earlier version of this script had an explicit O(rows x cols) fill-clear loop
    # here that was pure redundant work - removed since it noticeably slowed down
    # runs on the full day's data.
    sort_active_sheet(ws_active)
    apply_column_display(ws_active)

    active_rows = []
    header = ACTIVE_FIELDS
    for row in ws_active.iter_rows(min_row=2, values_only=True):
        active_rows.append(dict(zip(header, row)))

    build_aggregate(wb, active_rows)
    build_top10_sheet(wb, ws_active)

    wb.save(master_path)
    days_covered = ", ".join(d for d, _ in day_files)
    print(f"[{feed['label']}] processed day-files: {days_covered} | +{added} new, "
          f"-{removed} expired-pruned, -{neutral_removed} neutral-ticker-pruned "
          f"({len(neutral_tickers)} tickers), {skipped_expired} already-expired skipped, "
          f"{skipped_neutral} pre-skipped (already-neutral ticker) "
          f"-> {master_path} ({ws_active.max_row - 1} active rows)")
    if neutral_tickers:
        print(f"  [neutral] removed tickers (bullish~=bearish, within "
              f"{int(SKEW_NEUTRAL_BAND*100)}% band): {', '.join(sorted(neutral_tickers))}")
    if unparsed_strategies:
        print(f"  [warn] {sum(unparsed_strategies.values())} row(s) with unrecognized "
              f"strategy text (kept in Active, excluded from Aggregate). Distinct patterns:")
        for strat, count in sorted(unparsed_strategies.items(), key=lambda kv: -kv[1])[:10]:
            print(f"    x{count}: {strat}")

    # Only delete source day-files AFTER the master workbook has been saved
    # successfully, so a crash mid-run never loses data.
    deleted = cleanup_old_day_files(feed, day_files, KEEP_DAYS, today_date)
    if deleted:
        print(f"  [cleanup] deleted {len(deleted)} day-file(s) older than {KEEP_DAYS} days "
              f"(already merged): {', '.join(deleted)}")


def main():
    today_date = datetime.now().date()
    for feed in FEEDS:
        process_feed(feed, today_date)


if __name__ == "__main__":
    main()
